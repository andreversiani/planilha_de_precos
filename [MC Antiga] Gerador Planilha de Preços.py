from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import os

def get_mc_name():
  files = os.listdir()
  for file in files:
    if '- MC -' in file:
      mc_name = file
    
    if 'Planilha de Preço' in file and '- PC - ' in file:
      planilha_preco_name = file

  return planilha_preco_name, mc_name


TITLES = {
  'ENGENHARIA': {
  'white_list': ['Projetos']
  },

  'CIVIL': {
    'white_list': ['Obras Civis', 'Canteiro / Mobilização']
  },

  'MONTAGEM': {
    'white_list': ['Montagem Eletromecânica', 'Materiais', 'Eletrocentro']
  },

  'SERVIÇOS GERAIS': {
    'white_list': ['Treinamento', 'Comissionamento', 'Supervisão de Montagem', 'Administração de Obra', 'Frete', 'Despesas de Viagem']
  }
}

#sheet names
planilha_preco_sheet_name = 'Planilha de Preço'
planilha_preco_base_sheet_name = 'Base Planilha de Preço'
planilha_resumo_sheet_name = 'Planilha Resumo'
boilerplate_sheet_name = 'Boilerplate'
styles_sheet_name = 'Styles'
memo_sheet_name = 'Memo Geral'
db_sheet_name = 'DashBoard'

#colunas
conferencia_column = 'X'
descricao_column = 'B'
subestacao_column = 'Z'
qte_column = 'D'
preco_impostos_column = 'T'
erase_columns = ['E', 'G', 'I', 'K', 'L', 'N', 'O', 'P']
type_column = 'L'

#impostos
pis_cofins_eq = '$M$14'
pis_cofins_sv = '$M$13'
icms = '$M$22'
iss_bh = '$M$15'
iss_cliente = '$M$16'
ipi = "$M$12"

#cores
azul_escuro = "FFC5D9F1"
azul_claro = 'FFDAEEF3'
laranja = "FFFDE9D9"
branco = "FFFFFFFF"
roxo = "FFE4DFEC"

titles = ["ENGENHARIA", "ELÉTRICA", "CIVIL", "MONTAGEM", "SERVIÇOS GERAIS"]

TAXES = {
  'PR': {
    'pis_cofins': pis_cofins_eq,
    'icms': 0,
    'iss': iss_bh,
    'ipi': ipi,
  },
  'EQ': {
    'pis_cofins': pis_cofins_eq,
    'icms': icms,
    'iss': 0,
    'ipi': ipi,
  },
  'SV': {
    'pis_cofins': pis_cofins_sv,
    'icms': 0,
    'iss': iss_bh,
    'ipi': ipi
  },
  'SO': {
    'pis_cofins': pis_cofins_sv,
    'icms': 0,
    'iss': iss_bh,
    'ipi': ipi
  }
}


try:
  planilha_preco_name, mc_name = get_mc_name()
  wb_planilha_preco = load_workbook(planilha_preco_name, data_only=True)
  wb_mc = load_workbook(mc_name, data_only=True)

  planilha_preco_sheet_names = wb_planilha_preco.sheetnames
  planilha_preco_base = wb_planilha_preco[planilha_preco_base_sheet_name]
  
  planilha_preco = wb_planilha_preco.copy_worksheet(planilha_preco_base)

  if planilha_preco_sheet_name in planilha_preco_sheet_names:
    wb_planilha_preco.remove(wb_planilha_preco.worksheets[planilha_preco_sheet_names.index(planilha_preco_sheet_name)])

  if planilha_resumo_sheet_name in planilha_preco_sheet_names:
    wb_planilha_preco.remove(wb_planilha_preco.worksheets[planilha_preco_sheet_names.index(planilha_resumo_sheet_name)])

  planilha_preco.title = 'Planilha de Preço'
  planilha_preco.sheet_view.showGridLines = False
  
  resumo = wb_planilha_preco.copy_worksheet(planilha_preco_base)
  resumo.title = 'Planilha Resumo'
  resumo.sheet_view.showGridLines = False
  
  styles = wb_planilha_preco[styles_sheet_name]
  memo = wb_mc[memo_sheet_name]
  db = wb_mc[db_sheet_name]

except Exception as error:
  print(str(error))
  exit()

def qtd_validador(row, se):
  if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
    return True
  else: 
    return False

def make_item(se, white_list, type):
  i = 0
  se_count = 0
  se_names = get_se_names()

  print(f'{se.upper()} | Escrevendo a parte de {type}')
  
  for planilha_preco_row in range(8, 1000):
    title = planilha_preco[f'B{planilha_preco_row}'].value
    if title in se_names:
      se_count += 1
    index = se_names.index(se) + 1 

    if title == type and se_count == index:
      for memo_row in range(1, memo.max_row + 1):
        conferencia_value = memo[f'{conferencia_column}{memo_row}'].value
        if conferencia_value in white_list and qtd_validador(memo_row, se):
          taxes_type = memo[f'{type_column}{memo_row}'].value
          i += 1
          taxes = TAXES[taxes_type]
          complete_cells(planilha_preco_row, memo_row, i, taxes)


def get_se_row(se):
  for row in range(1, planilha_preco.max_row):
    if planilha_preco[f'B{row}'].value == se:
      return row

def complete_cells(planilha_preco_row, row, i, taxes):
  planilha_preco.insert_rows(planilha_preco_row + i, 1)
  #print(planilha_preco_row)
  #Somas
  planilha_preco.cell(row=planilha_preco_row, column=16, value=f"=SUM(P{planilha_preco_row+1}:P{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=15, value=f"=SUM(O{planilha_preco_row+1}:O{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=14, value=f"=SUM(N{planilha_preco_row+1}:N{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=12, value=f"=SUM(L{planilha_preco_row+1}:L{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=11, value=f"=SUM(K{planilha_preco_row+1}:K{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=9, value=f"=SUM(I{planilha_preco_row+1}:I{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=7, value=f"=SUM(G{planilha_preco_row+1}:G{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=5, value=f"=SUM(E{planilha_preco_row+1}:E{planilha_preco_row+i})")
      
  #Campos brancos
  planilha_preco.cell(row=planilha_preco_row + i, column=2, value=f"='[{mc_name}]Memo Geral'!${descricao_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=3, value=f"='[{mc_name}]Memo Geral'!${qte_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=4, value="R$")
  planilha_preco.cell(row=planilha_preco_row + i, column=16, value=f"='[{mc_name}]Memo Geral'!{preco_impostos_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=5, value=f"=L{planilha_preco_row + i}-K{planilha_preco_row + i}-I{planilha_preco_row + i}-G{planilha_preco_row + i}")
  planilha_preco.cell(row=planilha_preco_row + i, column=7, value=f"=L{planilha_preco_row + i}*F{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=9, value=f"=L{planilha_preco_row + i}*H{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=11, value=f"=L{planilha_preco_row + i}*J{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=12, value=f"=P{planilha_preco_row + i}/(1+M{planilha_preco_row + i}/100)")
  planilha_preco.cell(row=planilha_preco_row + i, column=14, value=f"=L{planilha_preco_row + i}*M{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=15, value=f"=G{planilha_preco_row + i}+I{planilha_preco_row + i}+K{planilha_preco_row + i}+N{planilha_preco_row + i}")
  
  #Impostos
  _icms = taxes['icms']
  _pis_cofins = taxes['pis_cofins']
  _iss = taxes['iss']
  _ipi = taxes['ipi']

  if _icms == 0:
    planilha_preco.cell(row=planilha_preco_row + i, column=8, value=0) #icms
  else:
    planilha_preco.cell(row=planilha_preco_row + i, column=8, value=f"='[{mc_name}]DashBoard'!{_icms} * 100") #icms
  if _iss == 0:
    planilha_preco.cell(row=planilha_preco_row + i, column=10, value=0) #iss
  else:
    planilha_preco.cell(row=planilha_preco_row + i, column=10, value=f"='[{mc_name}]DashBoard'!{_iss}") #iss
    
  planilha_preco.cell(row=planilha_preco_row + i, column=6, value=f"='[{mc_name}]DashBoard'!{_pis_cofins}") #pis/confins 
  planilha_preco.cell(row=planilha_preco_row + i, column=13, value=f"='[{mc_name}]DashBoard'!{_ipi}") #ipi
  
  #styles
  for column in range(1, planilha_preco.max_column + 1):
    cell = planilha_preco[f'{get_column_letter(column)}{planilha_preco_row + i}']
    cell._style = copy(styles[f'{get_column_letter(column)}3']._style)


def get_se_names():
  names_se = []
  for row in range(1, memo.max_row):
    value = memo[f'Z{row}'].internal_value
    if value != None and value[:2] == 'SE' and value not in names_se:
      names_se.append(value)
  return names_se

def get_other_titles_names():
  titles = []
  ignore = ['TOTAL/TITULO/VAZIO', 'Subestação / LT']
  for row in range(1, memo.max_row):
    value = memo[f'Z{row}'].internal_value
    if value != None and value[:2] != 'SE' and value not in titles and value not in ignore:
      num = memo[f'D{row}'].internal_value
      if num != None and int(num) > 0:
        titles.append(value)
  return titles

def get_total_row():
  for row in range(1, planilha_preco.max_row + 1):
    cell = f'A{row}'
    if planilha_preco[cell].value == 'TOTAL':
      return row

def make_titles(names_se):
  print('Fazendo os títulos')
  for se in names_se:
    
    if names_se.index(se) == 0: #preencher caso for a primeira SE
      planilha_preco.cell(row=4, column=1,value=str(se)) # Nome do título
      planilha_preco.cell(row=8, column=2, value=str(se)) # Nome da primeira SE
      planilha_preco.cell(row=8, column=1, value=int(names_se.index(se) + 1))
      
    else: #preenche o restante das SEs
      total_row = get_total_row()
      planilha_preco.insert_rows(total_row, 6)
      
      for row in range(total_row, total_row + 6):
        for column in range(1, planilha_preco.max_column + 1):
          
          copy_cell = planilha_preco[f'{get_column_letter(column)}{row-6}']
          new_cell = planilha_preco.cell(row=row, column=column, value="")
          new_cell._style = copy(copy_cell._style)

          if row == total_row and column == 2: # preenche o nome da SE
            planilha_preco.cell(row=row, column=column, value=str(se))

          elif column == 2:
            new_cell = planilha_preco.cell(row=row, column=column, value=copy_cell.value) #preenche a descrição

          if row == total_row and column == 1:
            planilha_preco.cell(row=row, column=column, value=int(names_se.index(se) + 1)) #preenche o primeiro item
        

def make_equipamentos(se):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Demais equipamentos de pátio" or cell.value == "Transformador de Força" or cell.value == "GIS / Módulo Híbrido":
      if memo[f'{subestacao_column}{memo_row}'].value == se and int(memo[f'{qte_column}{memo_row}'].value) > 0:
        exit = False
  if exit:
    return

  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1
  
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ELÉTRICA' and se_count == se_names.index(se) + 1:
      i = 1
      planilha_preco.insert_rows(planilha_preco_row + 1, 1)
      
      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}4']
        new_cell = planilha_preco.cell(row=planilha_preco_row + 1, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)

      planilha_preco.cell(row=planilha_preco_row + 1, column=2, value="EQUIPAMENTOS DE PÁTIO")

      for row in range(1, memo.max_row + 1):
        cell = memo[f'{conferencia_column}{row}']
        if cell.value == "Demais equipamentos de pátio" or cell.value == "GIS / Módulo Híbrido" or cell.value == "Transformador de Força":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            taxes_type = memo[f'{type_column}{row}'].value
            taxes = TAXES[taxes_type]
            complete_cells(planilha_preco_row + 1, row, i, taxes)
            i += 1

def make_casa(se):
  exit = True 
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Cubículos" or cell.value == "Proteção, medição e controle" or cell.value == "Telecomunicações":
      if memo[f'{subestacao_column}{memo_row}'].value == se and memo[f'{qte_column}{memo_row}'].value >= 1:
        exit = False
  if exit:
    return 0

  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL':
      se_count += 1
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL' and se_count == se_names.index(se) + 1:
      planilha_preco.insert_rows(planilha_preco_row, 1)

      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}4']
        new_cell = planilha_preco.cell(row=planilha_preco_row, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)
      
      planilha_preco.cell(row=planilha_preco_row, column=2, value="CASA DE COMANDO")
    
  make_casa_itens(se, "Cubículos", "CUBÍCULOS DE MÉDIA TENSÃO")
  make_casa_itens(se, "Proteção, medição e controle", "PROTEÇÃO, MEDIÇÃO E CONTROLE")
  make_casa_itens(se, "Telecomunicações", "TELECOMUNICAÇÕES")
  make_casa_itens(se, "Serviços Auxiliares", "SERVIÇOS AUXILIARES")

def make_casa_itens(se, memo_value, planilha_precos_title):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == memo_value:
      if memo[f'{subestacao_column}{memo_row}'].value == se and int(memo[f'{qte_column}{memo_row}'].value) > 0:
        exit = False
  if exit:
    return
  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL':
      se_count += 1
  
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL' and se_count == se_names.index(se) + 1:
      i = 1
      planilha_preco.insert_rows(planilha_preco_row, 1)
      
      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}5']
        new_cell = planilha_preco.cell(row=planilha_preco_row, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)

      planilha_preco.cell(row=planilha_preco_row, column=2, value=planilha_precos_title)

      for row in range(1, memo.max_row + 1):
        cell = memo[f'{conferencia_column}{row}']
        if cell.value == memo_value:
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            taxes_type = memo[f'{type_column}{row}'].value
            taxes = TAXES[taxes_type]
            complete_cells(planilha_preco_row, row, i, taxes)
            i += 1

def make_eletrica(se):
  se_status = get_se_status(se)
  if se_status['equipamentos']:
    make_equipamentos(se)
  if se_status['casa_comando']:
    make_casa(se)

def make_total_sums():
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  total_row = get_total_row()
  for column in total_columns:
    formula = "=SUM("
    for row in range(1, total_row + 1):
      cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == laranja:
        coordenada = cell.coordinate
        formula += coordenada + ","
    formula = formula[:-1]
    formula += ")"
    planilha_preco[f'{get_column_letter(column)}{total_row}'].value = formula

def make_se_sums(se):
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  se_info = get_se_status(se)
  first_row = se_info['first_row']
  last_row = se_info['last_row']
  
  for column in total_columns:
    formula = "=SUM("
    for row in range(first_row, last_row + 1):
      cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == azul_escuro:
        coordenada = cell.coordinate
        formula += coordenada + ","
    formula = formula[:-1]
    formula += ")"
    
    planilha_preco[f'{get_column_letter(column)}{first_row}'].value = formula

def make_eletrica_sums(se):
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  formula = False 
  
  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'ELÉTRICA':
      eletrica_row = row
    if cell == 'EQUIPAMENTOS DE PÁTIO':
      equipamentos_row = row
    if cell == 'CASA DE COMANDO':
      casa_comando_row = row

  for column in total_columns:
    column_letter = get_column_letter(column)
    
    if se_info['equipamentos'] and se_info['casa_comando']:
      formula = f'=SUM({column_letter}{equipamentos_row},{column_letter}{casa_comando_row})'
    
    if se_info['equipamentos'] and not se_info['casa_comando']:
      formula = f'=SUM({column_letter}{equipamentos_row})'
    
    if not se_info['equipamentos'] and se_info['casa_comando']:
      formula = f'=SUM({column_letter}{casa_comando_row})'
    
    if not formula:
      formula  = 0

    cell  = planilha_preco[f'{column_letter}{eletrica_row}']
    cell.value = formula

def make_casa_comando_sums(se):
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  casa_comando_row = 0
  
  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'CASA DE COMANDO':
      casa_comando_row = row   
  
  for column in total_columns:
    formula = "=SUM("
    for row in range(casa_comando_row, se_last_row + 1):
      cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == roxo:
        coordenada = cell.coordinate
        formula += coordenada + ","
    formula = formula[:-1]
    formula += ")"
    
    planilha_preco[f'{get_column_letter(column)}{casa_comando_row}'].value = formula

def make_indice_inicials(se):
  make_dark_blue_indice_inicials(se)
  make_light_blue_indice_inicials(se)
  make_purple_indice_inicials(se)
  make_white_indice_inicials(se)

def make_dark_blue_indice_inicials(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  index = 1
  
  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'A{row}']
    cell_color = cell.fill.start_color.rgb
    if cell_color == azul_escuro:
      formula = f'=A{se_first_row} & ".{index}"'
      cell.value = formula
      index += 1

def make_light_blue_indice_inicials(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  eletrica_row = 0
  index = 1

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'ELÉTRICA':
      eletrica_row = row   

  if eletrica_row:
    for row in range(eletrica_row, se_last_row + 1):
      cell = planilha_preco[f'A{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == azul_claro:
        formula = f'=A{eletrica_row} & ".{index}"'
        cell.value = formula
        index+= 1

def make_purple_indice_inicials(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  casa_comando_row = 0
  index = 1

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'CASA DE COMANDO':
      casa_comando_row = row   

  if casa_comando_row:
    for row in range(casa_comando_row, se_last_row + 1):
      cell = planilha_preco[f'A{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == roxo:
        formula = f'=A{casa_comando_row} & ".{index}"'
        cell.value = formula
        index+= 1

def make_white_indice_inicials(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  fixed_cell = 0

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'A{row}']
    cell_abaixo = planilha_preco[f'A{row + 1}']
    cell_color = cell.fill.start_color.rgb
    cell_abaixo_color = cell_abaixo.fill.start_color.rgb
    
    if cell_color != branco and cell_abaixo_color == branco:
      fixed_cell = cell
      formula = f'={fixed_cell.coordinate} & ".1"'
      cell_abaixo.value = formula

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'A{row}']
    cell_acima = planilha_preco[f'A{row - 1}']
    cell_color = cell.fill.start_color.rgb  
    
    if cell.value == None and cell_color == branco and cell_acima.value != None:
      texto = cell_acima.value
      index = int(texto[-2])
      formula = texto[:-2]
      formula += str(index + 1) + '"'
      cell.value = formula

def make_resumo():
  total_row = get_total_row()
  for row in range(8, total_row + 1):
    for column in range(1, planilha_preco.max_column + 1):
      copy_cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell = resumo[f'{get_column_letter(column)}{row}']
      cell._style = copy(copy_cell._style)
      title = planilha_preco[f'B{row}'].value
      title_index = planilha_preco[f'A{row}'].value

      if get_column_letter(column) not in erase_columns:
        cell.value = f"='{planilha_preco_sheet_name}'!{get_column_letter(column)}{row}"

      if title in titles or title_index in range(1, 101) or row == total_row:
        cell.value = f"='{planilha_preco_sheet_name}'!{get_column_letter(column)}{row}"

  for row in range(8, total_row + 1):
    erase_cell = resumo[f'{qte_column}{row}']
    erase_cell_color = erase_cell.fill.start_color.rgb
    if erase_cell_color != branco:
      resumo.cell(row=row, column=3, value="")
      resumo.cell(row=row, column=4, value="")
      resumo.cell(row=row, column=6, value="")
      resumo.cell(row=row, column=8, value="")
      resumo.cell(row=row, column=10, value="")
      resumo.cell(row=row, column=13, value="")

  resumo.cell(row=total_row, column=2, value="")
  resumo.cell(row=2, column=1, value=f"='{planilha_preco_sheet_name}'!A2")
  resumo.cell(row=4, column=1, value=f"='{planilha_preco_sheet_name}'!A4")

def make_other_titles():

  titles = get_other_titles_names()

  for title in titles:
    
    indice_inicial = len(get_se_names())
    total_row = get_total_row()
    planilha_preco.insert_rows(total_row, 1)
    title_row = total_row

    #TÍTULO
    for planilha_preco_column in range(1, planilha_preco.max_column + 1):
      indice = indice_inicial + titles.index(title) + 1
      copy_cell = styles[f'{get_column_letter(planilha_preco_column)}1']
      new_cell = planilha_preco.cell(row=title_row, column=planilha_preco_column, value=None)
      new_cell._style = copy(copy_cell._style)
      planilha_preco.cell(row=title_row, column=2, value=title)
      planilha_preco.cell(row=title_row, column=1, value=indice)

    #LINHAS EM BRANCO
    i = 1
    for memo_row in range(1, memo.max_column):
      conferencia_cell = memo[f'Z{memo_row}']
      qte_cell = memo[f'{qte_column}{memo_row}']
  
      if conferencia_cell.value == title and qte_cell.value != None and qte_cell.value > 0:
        taxes_type = memo[f'{type_column}{memo_row}'].value
        taxes = TAXES[taxes_type]
        complete_cells(title_row, memo_row, i, taxes)
        i += 1
    #SOMAS
  
    total_row = get_total_row()
    total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
    for column in total_columns:
      column_letter = get_column_letter(column)
      formula = f'=SUM({column_letter}{title_row + 1}:{column_letter}{total_row - 1})'
      cell = planilha_preco[f'{column_letter}{title_row}']
      cell.value = formula

    #ÍNDICE
    i = 1
    for row in range(title_row + 1, total_row):
      cell = planilha_preco[f'A{row}']
      cell.value = f'=A{title_row} & ".{i}"'

def get_se_status(se):
  se_status = {
    "name": se,
    "first_row": 0,
    "last_row": 0,
    "casa_comando": False,
    "equipamentos": False
  }
  
  total_row = get_total_row()
  for row in range(8, total_row):
    title = planilha_preco[f'B{row}'].value
    
    if title[:3] == "SE " and se_status["first_row"] > 0 and se_status["last_row"] == 0:
      se_status["last_row"] = row - 1

    if title == se:
      se_status["first_row"] = row

    if row == total_row - 1 and se_status['last_row'] == 0:
      se_status["last_row"] = total_row - 1

  for memo_row in range(1, memo.max_row):

    cell = memo[f'{conferencia_column}{memo_row}']
    qte_cell  = memo[f'{qte_column}{memo_row}']
    se_name_cell = memo[f'{subestacao_column}{memo_row}']

    if cell.value == "Cubículos" or cell.value == "Proteção, medição e controle" or cell.value == "Telecomunicações":
        if qte_cell.value > 0 and se_name_cell.value == se:
          se_status['casa_comando'] = True
          
  
    if cell.value == "Demais equipamentos de pátio" or cell.value == "Transformador de Força" or cell.value == "GIS / Módulo Híbrido":
      if qte_cell.value > 0 and se_name_cell.value == se:
        se_status['equipamentos'] = True

  return se_status

def make_se_names_header():
  se_names = get_se_names()
  cell = planilha_preco['A4']
  title_cell = planilha_preco['A2']
  title_cell.value = planilha_preco_name[:-5]
  
  if len(se_names) == 1:
    cell.value = str(se_names[0])
  else:
    header = se_names[:-1]
    header = (', ').join(header)
    header += ' E ' + se_names[-1]
    cell.value = header

def make_sums(se):
  se_status = get_se_status(se)
  make_se_sums(se)
  make_eletrica_sums(se)
  if se_status['casa_comando']:
    make_casa_comando_sums(se)

ENGENHARIA = {
  'white_list': ['Projetos']
}

CIVIL = {
  'white_list': ['Obras Civis', 'Canteiro / Mobilização']
}

MONTAGEM = {
  'white_list': ['Montagem Eletromecânica', 'Materiais', 'Eletrocentro']
}

MONTAGEM = {
  'white_list': ['Treinamento', 'Comissionamento', 'Supervisão de Montagem', 'Administração de Obra', 'Frete', 'Despesas de Viagem']
}


def build():
  se_names = get_se_names()
  make_titles(se_names)
  make_se_names_header()
  for se in se_names:
    make_item(se, TITLES['ENGENHARIA']['white_list'], list(TITLES)[0])
    make_eletrica(se)
    make_item(se, TITLES['CIVIL']['white_list'], list(TITLES)[1])
    make_item(se, TITLES['MONTAGEM']['white_list'], list(TITLES)[2])
    make_item(se, TITLES['SERVIÇOS GERAIS']['white_list'], list(TITLES)[3])
    make_sums(se)
    make_indice_inicials(se)
  
  make_other_titles()
  make_total_sums()
  make_resumo()
  
  try:
    wb_planilha_preco.save(planilha_preco_name)
  except Exception as error:
    print("Feche a MC antes de rodar o Gerador de Planilha de Preços")

build()