from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import os

def get_mc_name():
  folder_path = os.getcwd()
  folder_path_array = folder_path.split('\\')
  numero_proposta = folder_path_array[-1]
  nome_proposta = folder_path_array[-3]
  nomes = nome_proposta.split(' ')
  nomes.insert(2, "MC -")
  nomes = nomes[1:]
  nomes.insert(0, numero_proposta)
  nome_mc = ' '.join(nomes) + '.xlsm'
  planilha_preco_name = f'{numero_proposta} - PC - ANEXO 01 - Planilha de Preço.xlsx'
  return planilha_preco_name, nome_mc

#sheet names
planilha_preco_sheet_name = 'Planilha de Preço'
planilha_resumo_sheet_name = 'Planilha Resumo'
boilerplate_sheet_name = 'Boilerplate'
styles_sheet_name = 'Styles'
memo_sheet_name = 'Memo Geral'
db_sheet_name = 'DashBoard'

#colunas
conferencia_column = 'X'
descricao_column = 'B'
subestacao_column = 'Z'
qte_column = 'D'
preco_impostos_column = 'T'
erase_columns = ['E', 'G', 'I', 'K', 'L', 'N', 'O', 'P'] 

#impostos
pis_confins_eq = '$M$14'
pis_confins_sv = '$M$13'
icms = '$M$22'
iss_bh = '$M$15'
iss_cliente = '$M$16'
ipi = "$M$12"

#cores
azul_escuro = "FFC5D9F1"
azul_claro = 'FFDAEEF3'
laranja = "FFFDE9D9"
branco = "FFFFFFFF"
roxo = "FFE4DFEC"

titles = ["ENGENHARIA", "ELÉTRICA", "CIVIL", "MONTAGEM", "SERVIÇOS GERAIS"]

try:
  planilha_preco_name, mc_name = get_mc_name()
  wb_planilha_preco = load_workbook(planilha_preco_name, data_only=True)
  wb_mc = load_workbook(mc_name, data_only=True)
  planilha_preco = wb_planilha_preco[planilha_preco_sheet_name]
  resumo = wb_planilha_preco[planilha_resumo_sheet_name]
  styles = wb_planilha_preco[styles_sheet_name]
  boilerplate = wb_planilha_preco[boilerplate_sheet_name]
  memo = wb_mc[memo_sheet_name]
  db = wb_mc[db_sheet_name]
except Exception:
  print("Erro ao achar a MC")
  exit()

def get_se_row(se):
  for row in range(1, planilha_preco.max_row):
    if planilha_preco[f'B{row}'].value == se:
      return row

def complete_cells(planilha_preco_row, row, i, eletrocentro=False):
  planilha_preco.insert_rows(planilha_preco_row + i, 1)
  #Somas
  planilha_preco.cell(row=planilha_preco_row, column=16, value=f"=SUM(P{planilha_preco_row+1}:P{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=15, value=f"=SUM(O{planilha_preco_row+1}:O{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=14, value=f"=SUM(N{planilha_preco_row+1}:N{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=12, value=f"=SUM(L{planilha_preco_row+1}:L{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=11, value=f"=SUM(K{planilha_preco_row+1}:K{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=9, value=f"=SUM(I{planilha_preco_row+1}:I{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=7, value=f"=SUM(G{planilha_preco_row+1}:G{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=5, value=f"=SUM(E{planilha_preco_row+1}:E{planilha_preco_row+i})")
      
  #Campos brancos
  planilha_preco.cell(row=planilha_preco_row + i, column=2, value=f"='[{mc_name}]Memo Geral'!${descricao_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=3, value=f"='[{mc_name}]Memo Geral'!${qte_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=4, value="R$")
  planilha_preco.cell(row=planilha_preco_row + i, column=16, value=f"='[{mc_name}]Memo Geral'!{preco_impostos_column}${row}")
  
  #formulas fixadas nos campos brancos
  if not eletrocentro:
    planilha_preco.cell(row=planilha_preco_row + i, column=6, value=f"=$F${planilha_preco_row+1}")
    planilha_preco.cell(row=planilha_preco_row + i, column=8, value=f"=$H${planilha_preco_row+1}")
    planilha_preco.cell(row=planilha_preco_row + i, column=10, value=f"=$J${planilha_preco_row+1}")
    planilha_preco.cell(row=planilha_preco_row + i, column=13, value=f"=$M${planilha_preco_row+1}")

  if eletrocentro:
    planilha_preco.cell(row=planilha_preco_row + i, column=6, value=f"='[{mc_name}]DashBoard'!{pis_confins_eq}")
    planilha_preco.cell(row=planilha_preco_row + i, column=8, value=f"='[{mc_name}]DashBoard'!{icms} * 100")
    planilha_preco.cell(row=planilha_preco_row + i, column=10, value=0)
    planilha_preco.cell(row=planilha_preco_row + i, column=13, value=0)
  
  planilha_preco.cell(row=planilha_preco_row + i, column=5, value=f"=L{planilha_preco_row + i}-K{planilha_preco_row + i}-I{planilha_preco_row + i}-G{planilha_preco_row + i}")
  planilha_preco.cell(row=planilha_preco_row + i, column=7, value=f"=L{planilha_preco_row + i}*F{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=9, value=f"=L{planilha_preco_row + i}*H{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=11, value=f"=L{planilha_preco_row + i}*J{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=12, value=f"=P{planilha_preco_row + i}/(1+{planilha_preco_row + i}/100)")
  planilha_preco.cell(row=planilha_preco_row + i, column=14, value=f"=L{planilha_preco_row + i}*M{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=15, value=f"=G{planilha_preco_row + i}+I{planilha_preco_row + i}+K{planilha_preco_row + i}+N{planilha_preco_row + i}")

  #styles
  for column in range(1, planilha_preco.max_column + 1):
    cell = planilha_preco[f'{get_column_letter(column)}{planilha_preco_row + i}']
    cell._style = copy(styles[f'{get_column_letter(column)}3']._style)

def make_taxes(se, subtopico, pis_confins, icms, iss, ipi, sobressalente_row = None):
  if subtopico != "SOBRESSALENTES":
    row = 1
    while planilha_preco[f'B{row}'].value != se:
      row += 1
    while planilha_preco[f'B{row}'].value != subtopico:
      row += 1
    row += 1
  
  if subtopico == "SOBRESSALENTES":
    row = sobressalente_row

  if icms == 0:
    planilha_preco.cell(row=row, column=8, value=0) #icms
  else:
    planilha_preco.cell(row=row, column=8, value=f"='[{mc_name}]DashBoard'!{icms} * 100") #icms
  if iss == 0:
    planilha_preco.cell(row=row, column=10, value=0) #iss
  else:
    planilha_preco.cell(row=row, column=10, value=f"='[{mc_name}]DashBoard'!{iss}") #iss
    
  planilha_preco.cell(row=row, column=6, value=f"='[{mc_name}]DashBoard'!{pis_confins}") #pis/confins 
  planilha_preco.cell(row=row, column=13, value=f"='[{mc_name}]DashBoard'!{ipi}") #ipi


def get_se_names():
  names_se = []
  for row in range(1, memo.max_row):
    value = memo[f'Z{row}'].internal_value
    if value != None and value[:2] == 'SE' and value not in names_se:
      names_se.append(value)
  return names_se

def get_total_row():
  for row in range(1, planilha_preco.max_row + 1):
    cell = f'A{row}'
    if planilha_preco[cell].value == 'TOTAL':
      return row

def make_titles(names_se):
  print('Fazendo os títulos')
  for se in names_se:
    
    if names_se.index(se) == 0: #preencher caso for a primeira SE
      planilha_preco.cell(row=4, column=1,value=str(se)) # Nome do título
      planilha_preco.cell(row=8, column=2, value=str(se)) # Nome da primeira SE
      planilha_preco.cell(row=8, column=1, value=int(names_se.index(se) + 1))
      
    else: #preenche o restante das SEs
      total_row = get_total_row()
      planilha_preco.insert_rows(total_row, 6)
      
      for row in range(total_row, total_row + 6):
        for column in range(1, planilha_preco.max_column + 1):
          
          copy_cell = planilha_preco[f'{get_column_letter(column)}{row-6}']
          new_cell = planilha_preco.cell(row=row, column=column, value="")
          new_cell._style = copy(copy_cell._style)

          if row == total_row and column == 2: # preenche o nome da SE
            planilha_preco.cell(row=row, column=column, value=str(se))

          elif column == 2:
            new_cell = planilha_preco.cell(row=row, column=column, value=copy_cell.value) #preenche a descrição

          if row == total_row and column == 1:
            planilha_preco.cell(row=row, column=column, value=int(names_se.index(se) + 1)) #preenche o primeiro item
        
def make_engenharia(se):
  print(f'{se.upper()} | Escrevendo a parte de Engenharia')
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA' and planilha_preco[f'B{planilha_preco_row-1}'].value == se:
      #se = str(planilha_preco[f'B{planilha_preco_row - 1}'].value)
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Projetos" and memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
          complete_cells(planilha_preco_row, row, i)
          i += 1
  make_taxes(se=se, subtopico='ENGENHARIA', pis_confins=pis_confins_eq, icms=0, iss=iss_bh, ipi=ipi)

def make_civil(se, se_names):
  print(f'{se.upper()} | Escrevendo a parte de Civil')
  se_count=  0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1
    
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL'and se_count == se_names.index(se) + 1:
      #se = str(planilha_preco[engenharia_cell.coordinate].value)
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Obras Civis" or memo[f'{conferencia_column}{row}'].value == "Canteiro / Mobilização":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row, row, i)
            i += 1
  #impostos
  make_taxes(se=se, subtopico='CIVIL', pis_confins=pis_confins_sv, icms=0, iss=iss_cliente, ipi=ipi)


def make_montagem(se, se_names):
  print(f'{se.upper()} | Escrevendo a parte de Montagem')
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1

    if planilha_preco[f'B{planilha_preco_row}'].value == 'MONTAGEM' and se_count == se_names.index(se) + 1:
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Montagem Eletromecânica" or memo[f'{conferencia_column}{row}'].value == "Materiais" or memo[f'{conferencia_column}{row}'].value == "Eletrocentro":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            if memo[f'{conferencia_column}{row}'].value == "Eletrocentro":
              complete_cells(planilha_preco_row, row, i, eletrocentro=True)
            else:
              complete_cells(planilha_preco_row, row, i)
            i += 1
  #impostos
  make_taxes(se=se, subtopico='MONTAGEM', pis_confins=pis_confins_sv, icms=0, iss=iss_cliente, ipi=ipi)


def make_servicos_gerais(se):
  print(f'{se.upper()} | Escrevendo a parte de Serviços Gerais')
  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1

    if planilha_preco[f'B{planilha_preco_row}'].value == 'SERVIÇOS GERAIS' and se_count == se_names.index(se) + 1:
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Treinamento" or memo[f'{conferencia_column}{row}'].value == "Comissionamento" or memo[f'{conferencia_column}{row}'].value == "Supervisão de Montagem" or memo[f'{conferencia_column}{row}'].value == "Administração de Obra" or memo[f'{conferencia_column}{row}'].value == "Frete" or memo[f'{conferencia_column}{row}'].value == "Despesas de Viagem":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row, row, i)
            i += 1

  make_taxes(se=se, subtopico='SERVIÇOS GERAIS', pis_confins=pis_confins_sv, icms=0, iss=iss_cliente, ipi=ipi)


def make_equipamentos(se):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Demais equipamentos de pátio" or cell.value == "Transformador de Força" or cell.value == "GIS / Módulo Híbrido":
      if memo[f'{subestacao_column}{memo_row}'].value == se and int(memo[f'{qte_column}{memo_row}'].value) > 0:
        exit = False
  if exit:
    return

  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1
  
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ELÉTRICA' and se_count == se_names.index(se) + 1:
      i = 1
      planilha_preco.insert_rows(planilha_preco_row + 1, 1)
      
      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}4']
        new_cell = planilha_preco.cell(row=planilha_preco_row + 1, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)

      planilha_preco.cell(row=planilha_preco_row + 1, column=2, value="EQUIPAMENTOS DE PÁTIO")

      for row in range(1, memo.max_row + 1):
        cell = memo[f'{conferencia_column}{row}']
        if cell.value == "Demais equipamentos de pátio" or cell.value == "GIS / Módulo Híbrido" or cell.value == "Transformador de Força":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row + 1, row, i)
            i += 1
  
  make_taxes(se=se, subtopico='EQUIPAMENTOS DE PÁTIO', pis_confins=pis_confins_eq, icms=icms, iss=0, ipi=ipi)

def make_casa(se):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Cubículos" or cell.value == "Proteção, medição e controle" or cell.value == "Telecomunicações":
      if memo[f'{subestacao_column}{memo_row}'].value == se and memo[f'{qte_column}{memo_row}'].value >= 1:
        exit = False
  if exit:
    return 0

  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL':
      se_count += 1
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL' and se_count == se_names.index(se) + 1:
      planilha_preco.insert_rows(planilha_preco_row, 1)

      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}4']
        new_cell = planilha_preco.cell(row=planilha_preco_row, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)
      
      planilha_preco.cell(row=planilha_preco_row, column=2, value="CASA DE COMANDO")
    
  make_casa_itens(se, "Cubículos", "CUBÍCULOS DE MÉDIA TENSÃO")
  make_casa_itens(se, "Proteção, medição e controle", "PROTEÇÃO, MEDIÇÃO E CONTROLE")
  make_casa_itens(se, "Telecomunicações", "TELECOMUNICAÇÕES")
  make_casa_itens(se, "Serviços Auxiliares", "SERVIÇOS AUXILIARES")

def make_casa_itens(se, memo_value, planilha_precos_title):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == memo_value:
      if memo[f'{subestacao_column}{memo_row}'].value == se and int(memo[f'{qte_column}{memo_row}'].value) > 0:
        exit = False
  if exit:
    return
  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL':
      se_count += 1
  
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL' and se_count == se_names.index(se) + 1:
      i = 1
      planilha_preco.insert_rows(planilha_preco_row, 1)
      
      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}5']
        new_cell = planilha_preco.cell(row=planilha_preco_row, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)

      planilha_preco.cell(row=planilha_preco_row, column=2, value=planilha_precos_title)

      for row in range(1, memo.max_row + 1):
        cell = memo[f'{conferencia_column}{row}']
        if cell.value == memo_value:
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row, row, i)
            i += 1
  
  make_taxes(se=se, subtopico=planilha_precos_title, pis_confins=pis_confins_eq, icms=icms, iss=0, ipi=ipi)

def make_eletrica(se):
  make_equipamentos(se)
  make_casa(se)

def make_total_sums():
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  total_row = get_total_row()
  for column in total_columns:
    formula = "=SUM("
    for row in range(1, total_row + 1):
      cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == laranja:
        coordenada = cell.coordinate
        formula += coordenada + ","
    formula = formula[:-1]
    formula += ")"
    planilha_preco[f'{get_column_letter(column)}{total_row}'].value = formula

def make_se_sums(se):
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  se_info = get_se_status(se)
  first_row = se_info['first_row']
  last_row = se_info['last_row']
  
  for column in total_columns:
    formula = "=SUM("
    for row in range(first_row, last_row + 1):
      cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == azul_escuro:
        coordenada = cell.coordinate
        formula += coordenada + ","
    formula = formula[:-1]
    formula += ")"
    
    planilha_preco[f'{get_column_letter(column)}{first_row}'].value = formula

def make_eletrica_sums(se):
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  
  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'ELÉTRICA':
      eletrica_row = row
    if cell == 'EQUIPAMENTOS DE PÁTIO':
      equipamentos_row = row
    if cell == 'CASA DE COMANDO':
      casa_comando_row = row   

  for column in total_columns:
    column_letter = get_column_letter(column)
    formula = f'=SUM({column_letter}{equipamentos_row}, {column_letter}{casa_comando_row})'
    cell  = planilha_preco[f'{column_letter}{eletrica_row}']
    cell.value = formula

def make_casa_comando_sums(se):
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  casa_comando_row = 0
  
  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'CASA DE COMANDO':
      casa_comando_row = row   
  
  for column in total_columns:
    formula = "=SUM("
    for row in range(casa_comando_row, se_last_row + 1):
      cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == roxo:
        coordenada = cell.coordinate
        formula += coordenada + ","
    formula = formula[:-1]
    formula += ")"
    
    planilha_preco[f'{get_column_letter(column)}{casa_comando_row}'].value = formula

def make_indices(se):
  make_dark_blue_indices(se)
  make_light_blue_indices(se)
  make_purple_indices(se)
  make_white_indices(se)

def make_dark_blue_indices(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  index = 1
  
  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'A{row}']
    cell_color = cell.fill.start_color.rgb
    if cell_color == azul_escuro:
      formula = f'=A{se_first_row} & ".{index}"'
      cell.value = formula
      index += 1

def make_light_blue_indices(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  eletrica_row = 0
  index = 1

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'ELÉTRICA':
      eletrica_row = row   

  if eletrica_row:
    for row in range(eletrica_row, se_last_row + 1):
      cell = planilha_preco[f'A{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == azul_claro:
        formula = f'=A{eletrica_row} & ".{index}"'
        cell.value = formula
        index+= 1

def make_purple_indices(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  casa_comando_row = 0
  index = 1

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'B{row}'].value
    if cell == 'CASA DE COMANDO':
      casa_comando_row = row   

  if casa_comando_row:
    for row in range(casa_comando_row, se_last_row + 1):
      cell = planilha_preco[f'A{row}']
      cell_color = cell.fill.start_color.rgb
      if cell_color == roxo:
        formula = f'=A{casa_comando_row} & ".{index}"'
        cell.value = formula
        index+= 1

def make_white_indices(se):
  se_info = get_se_status(se)
  se_first_row = se_info['first_row']
  se_last_row = se_info['last_row']
  fixed_cell = 0

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'A{row}']
    cell_abaixo = planilha_preco[f'A{row + 1}']
    cell_color = cell.fill.start_color.rgb
    cell_abaixo_color = cell_abaixo.fill.start_color.rgb
    
    if cell_color != branco and cell_abaixo_color == branco:
      fixed_cell = cell
      formula = f'={fixed_cell.coordinate} & ".1"'
      cell_abaixo.value = formula

  for row in range(se_first_row, se_last_row + 1):
    cell = planilha_preco[f'A{row}']
    cell_acima = planilha_preco[f'A{row - 1}']
    cell_color = cell.fill.start_color.rgb  
    
    if cell.value == None and cell_color == branco and cell_acima.value != None:
      texto = cell_acima.value
      index = int(texto[-2])
      formula = texto[:-2]
      formula += str(index + 1) + '"'
      cell.value = formula

def make_resumo():
  total_row = get_total_row()
  for row in range(8, total_row + 1):
    for column in range(1, planilha_preco.max_column + 1):
      copy_cell = planilha_preco[f'{get_column_letter(column)}{row}']
      cell = resumo[f'{get_column_letter(column)}{row}']
      cell._style = copy(copy_cell._style)
      title = planilha_preco[f'B{row}'].value
      title_index = planilha_preco[f'A{row}'].value

      if get_column_letter(column) not in erase_columns:
        cell.value = f"='{planilha_preco_sheet_name}'!{get_column_letter(column)}{row}"

      if title in titles or title_index in range(1, 101) or row == total_row:
        cell.value = f"='{planilha_preco_sheet_name}'!{get_column_letter(column)}{row}"

  for row in range(8, total_row + 1):
    erase_cell = resumo[f'{qte_column}{row}']
    erase_cell_color = erase_cell.fill.start_color.rgb
    if erase_cell_color != branco:
      resumo.cell(row=row, column=3, value="")
      resumo.cell(row=row, column=4, value="")
      resumo.cell(row=row, column=6, value="")
      resumo.cell(row=row, column=8, value="")
      resumo.cell(row=row, column=10, value="")
      resumo.cell(row=row, column=13, value="")

  resumo.cell(row=total_row, column=2, value="")
  resumo.cell(row=2, column=1, value=f"='{planilha_preco_sheet_name}'!A2")
  resumo.cell(row=4, column=1, value=f"='{planilha_preco_sheet_name}'!A4")

def make_sobressalentes():
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Sobressalentes" and int(memo[f'{qte_column}{memo_row}'].value):
        exit = False
  if exit:
    return 0
  
  indice = len(get_se_names()) + 1
  total_row = get_total_row()
  planilha_preco.insert_rows(total_row, 1)
  sobressalente_row = total_row

  #TÍTULO
  for planilha_preco_column in range(1, planilha_preco.max_column + 1):
    copy_cell = styles[f'{get_column_letter(planilha_preco_column)}1']
    new_cell = planilha_preco.cell(row=sobressalente_row, column=planilha_preco_column, value=None)
    new_cell._style = copy(copy_cell._style)
    planilha_preco.cell(row=sobressalente_row, column=2, value='SOBRESSALENTES')
    planilha_preco.cell(row=sobressalente_row, column=1, value=indice)

  #LINHAS EM BRANCO
  i = 0
  for memo_row in range(1, memo.max_column):
    conferencia_cell = memo[f'{conferencia_column}{memo_row}']
    if conferencia_cell.value == "Sobressalentes":
      complete_cells(sobressalente_row + 1, memo_row, i)
      i += 1
  make_taxes(se='Sobressalentes', subtopico='SOBRESSALENTES', pis_confins=pis_confins_eq, icms=icms, iss=0, ipi=ipi, sobressalente_row=sobressalente_row+1)

  #SOMAS
  total_row = get_total_row()
  total_columns = [5, 7, 9, 11, 12, 14, 15, 16]
  for column in total_columns:
    column_letter = get_column_letter(column)
    formula = f'=SUM({column_letter}{sobressalente_row+1}:{column_letter}{total_row - 1})'
    cell = planilha_preco[f'{column_letter}{sobressalente_row}']
    cell.value = formula

  #ÍNDICE
  i = 1
  for row in range(sobressalente_row + 1, total_row):
    cell = planilha_preco[f'A{row}']
    cell.value = f'=A{sobressalente_row} & ".{i}"'

def get_se_status(se):
  se_status = {
    "name": se,
    "first_row": 0,
    "last_row": 0
  }
  total_row = get_total_row()
  for row in range(8, total_row):
    title = planilha_preco[f'B{row}'].value
    
    if title[:3] == "SE " and se_status["first_row"] > 0:
      se_status["last_row"] = row - 1

    if title == se:
      se_status["first_row"] = row

    if row == total_row - 1 and se_status['last_row'] == 0:
      se_status["last_row"] = total_row - 1
  
  return se_status

def make_se_names_header():
  se_names = get_se_names()
  cell = planilha_preco['A4']
  title_cell = planilha_preco['A2']
  title_cell.value = planilha_preco_name[:-5]
  
  if len(se_names) == 1:
    cell.value = str(se_names[0])
  else:
    header = se_names[:-1]
    header = (', ').join(header)
    header += ' E ' + se_names[-1]
    cell.value = header

def setup():
  
  source = boilerplate
  target = wb_planilha_preco.copy_worksheet(source)
  target.sheet_view.showGridLines = False

  
  #for row in range(8, 200):
    
    #for column in range(1, planilha_preco.max_column + 1):
      #column_letter = get_column_letter(column)
      #cell = planilha_preco[f'{column_letter}{row}']
      #copy_cell = boilerplate[f'A15']
      #cell.value = None
      #cell._style = copy(copy_cell)
  
  #for row in range(8, boilerplate.max_row + 1):
    #for column in range(1, boilerplate.max_column + 1):
      #column_letter = get_column_letter(column)
      #copy_cell = boilerplate[f'{column_letter}{row}']
      #cell = planilha_preco[f'{column_letter}{row}']
      #cell.value = copy_cell.value
      #cell._style = copy(copy_cell._style)

def build():
  se_names = get_se_names()
  setup()
  #make_titles(se_names)
  #make_se_names_header()
  #for se in se_names:
    #make_engenharia(se)
    #make_eletrica(se)
    #make_civil(se, se_names)
    #make_montagem(se, se_names)
    #make_servicos_gerais(se)
    #make_se_sums(se)
    #make_eletrica_sums(se)
    #make_casa_comando_sums(se)
    #make_indices(se)
  
  #make_sobressalentes()
  #make_total_sums()
  #make_resumo()
  wb_planilha_preco.save(planilha_preco_name)

build()