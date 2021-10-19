from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

planilha_preco_name = 'F4792-001-00 - PC - ANEXO 01 - Planilha de Preço.xlsx'
mc_name = 'F4792-001-00 - MC - PIAUÍ NÍQUEL - Bay de Saída 230 kV.xlsm'

wb_planilha_preco = load_workbook(planilha_preco_name)
wb_mc = load_workbook(mc_name, data_only=True)

planilha_preco = wb_planilha_preco['Teste']
styles = wb_planilha_preco['Styles']
memo = wb_mc['Memo Geral']
db = wb_mc['DashBoard']

conferencia_column = 'X'
descricao_column = 'B'
subestacao_column = 'Z'
qte_column = 'D'
preco_impostos_column = 'T'

pis_confins_eq = '$M$14'
pis_confins_sv = '$M$13'
icms = '$M$22'
iss_bh = '$M$15'
iss_cliente = '$M$16'
ipi = "$M$12"

def complete_cells(planilha_preco_row, row, i):
  planilha_preco.insert_rows(planilha_preco_row + i, 1)
  #Somas
  planilha_preco.cell(row=planilha_preco_row, column=16, value=f"=SUM(P{planilha_preco_row+1}:P{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=15, value=f"=SUM(O{planilha_preco_row+1}:O{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=14, value=f"=SUM(N{planilha_preco_row+1}:N{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=12, value=f"=SUM(L{planilha_preco_row+1}:L{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=11, value=f"=SUM(K{planilha_preco_row+1}:K{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=9, value=f"=SUM(I{planilha_preco_row+1}:I{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=7, value=f"=SUM(G{planilha_preco_row+1}:G{planilha_preco_row+i})")
  planilha_preco.cell(row=planilha_preco_row, column=5, value=f"=SUM(E{planilha_preco_row+1}:E{planilha_preco_row+i})")
      
  #Campos brancos
  planilha_preco.cell(row=planilha_preco_row + i, column=2, value=f"='[{mc_name}]Memo Geral'!${descricao_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=3, value=f"='[{mc_name}]Memo Geral'!${qte_column}${row}")
  planilha_preco.cell(row=planilha_preco_row + i, column=4, value="R$")
  planilha_preco.cell(row=planilha_preco_row + i, column=16, value=f"='[{mc_name}]Memo Geral'!{preco_impostos_column}${row}")
  
  #formulas fixadas nos campos brancos
  planilha_preco.cell(row=planilha_preco_row + i, column=6, value=f"=$F${planilha_preco_row+1}")
  planilha_preco.cell(row=planilha_preco_row + i, column=8, value=f"=$H${planilha_preco_row+1}")
  planilha_preco.cell(row=planilha_preco_row + i, column=10, value=f"=$J${planilha_preco_row+1}")
  planilha_preco.cell(row=planilha_preco_row + i, column=13, value=f"=$M${planilha_preco_row+1}")
  planilha_preco.cell(row=planilha_preco_row + i, column=5, value=f"=L{planilha_preco_row + i}-K{planilha_preco_row + i}-I{planilha_preco_row + i}-G{planilha_preco_row + i}")
  planilha_preco.cell(row=planilha_preco_row + i, column=7, value=f"=L{planilha_preco_row + i}*F{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=9, value=f"=L{planilha_preco_row + i}*H{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=11, value=f"=L{planilha_preco_row + i}*J{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=12, value=f"=P{planilha_preco_row + i}/(1+{planilha_preco_row + i}/100)")
  planilha_preco.cell(row=planilha_preco_row + i, column=14, value=f"=L{planilha_preco_row + i}*M{planilha_preco_row + i}/100")
  planilha_preco.cell(row=planilha_preco_row + i, column=15, value=f"=G{planilha_preco_row + i}+I{planilha_preco_row + i}+K{planilha_preco_row + i}+N{planilha_preco_row + i}")
  
  #indices
  #planilha_preco.cell(row=planilha_preco_row, column=1, value=f'=A{planilha_preco_row-1}&".1"')
  #planilha_preco.cell(row=planilha_preco_row + i, column=1, value=f'=A{planilha_preco_row-1}&".{i}"')
  
  #estilos
  for column in range(1, planilha_preco.max_column + 1):
    cell = planilha_preco[f'{get_column_letter(column)}{planilha_preco_row + i}']
    cell._style = copy(styles[f'{get_column_letter(column)}3']._style)

def make_taxes(se, subtopico, pis_confins, icms, iss, ipi):
  se_count = 0
  se_names = get_se_names()

  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value in se_names:
      se_count += 1

  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == subtopico and se_count == se_names.index(se) + 1:
      
      if icms == 0:
        planilha_preco.cell(row=planilha_preco_row + 1, column=8, value=0) #icms
      else:
        planilha_preco.cell(row=planilha_preco_row + 1, column=8, value=f"='[{mc_name}]DashBoard'!{icms} * 100") #icms
      if iss == 0:
        planilha_preco.cell(row=planilha_preco_row + 1, column=10, value=0) #iss
      else:
        planilha_preco.cell(row=planilha_preco_row + 1, column=10, value=f"='[{mc_name}]DashBoard'!{iss}") #iss
      
      planilha_preco.cell(row=planilha_preco_row + 1, column=6, value=f"='[{mc_name}]DashBoard'!{pis_confins}") #pis/confins 
      planilha_preco.cell(row=planilha_preco_row + 1, column=13, value=f"='[{mc_name}]DashBoard'!{ipi}") #ipi

def get_se_names():
  names_se = []
  for row in range(1, memo.max_row):
    value = memo[f'Z{row}'].internal_value
    if value != None and value[0] == 'S' and value[1] == 'E' and value not in names_se:
      names_se.append(value)
  return names_se

def get_total_row():
  for row in range(1, planilha_preco.max_row + 1):
    cell = f'A{row}'
    if planilha_preco[cell].value == 'TOTAL':
      return row

def make_titles(names_se):
  print('Fazendo os títulos')
  for se in names_se:
    if names_se.index(se) == 0: #preencher caso for a primeira SE
      planilha_preco.cell(row=4, column=1,value=str(se)) # Nome do título
      planilha_preco.cell(row=8, column=2, value=str(se)) # Nome da primeira SE
    
    else: #preenche o restante das SEs
      total_row = get_total_row()
      planilha_preco.insert_rows(total_row, 6)
      
      for row in range(total_row, total_row + 6):
        for column in range(1, planilha_preco.max_column + 1):
        
          copy_cell = planilha_preco[f'{get_column_letter(column)}{row-6}']
          new_cell = planilha_preco.cell(row=row, column=column, value="")
          new_cell._style = copy(copy_cell._style)

          if row == total_row and column == 2: # preenche o nome da SE
            planilha_preco.cell(row=row, column=column, value=str(se))

          elif column == 2:
            new_cell = planilha_preco.cell(row=row, column=column, value=copy_cell.value) #preenche a descrição

          if row == total_row and column == 1:
            planilha_preco.cell(row=row, column=column, value=int(names_se.index(se) + 1)) #preenche o primeiro item
          

def make_engenharia(se):
  print(f'{se.upper()} | Escrevendo a parte de Engenharia')
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA' and planilha_preco[f'B{planilha_preco_row-1}'].value == se:
      #se = str(planilha_preco[f'B{planilha_preco_row - 1}'].value)
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Projetos" and memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
          complete_cells(planilha_preco_row, row, i)
          i += 1
          #impostos
  make_taxes(se=se, subtopico='ENGENHARIA', pis_confins=pis_confins_eq, icms=0, iss=iss_bh, ipi=ipi)


def make_civil(se, se_names):
  print(f'{se.upper()} | Escrevendo a parte de Civil')
  se_count=  0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1
    
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL'and se_count == se_names.index(se) + 1:
      #se = str(planilha_preco[engenharia_cell.coordinate].value)
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Obras Civis" or memo[f'{conferencia_column}{row}'].value == "Canteiro / Mobilização":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row, row, i)
            i += 1
  #impostos
  make_taxes(se=se, subtopico='CIVIL', pis_confins=pis_confins_sv, icms=0, iss=iss_cliente, ipi=ipi)


def make_montagem(se, se_names):
  print(f'{se.upper()} | Escrevendo a parte de Montagem')
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1

    if planilha_preco[f'B{planilha_preco_row}'].value == 'MONTAGEM' and se_count == se_names.index(se) + 1:
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Montagem Eletromecânica" or memo[f'{conferencia_column}{row}'].value == "Materiais":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row, row, i)
            i += 1
  #impostos
  make_taxes(se=se, subtopico='MONTAGEM', pis_confins=pis_confins_sv, icms=0, iss=iss_cliente, ipi=ipi)

def make_servicos_gerais(se):
  print(f'{se.upper()} | Escrevendo a parte de Serviços Gerais')
  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1

    if planilha_preco[f'B{planilha_preco_row}'].value == 'SERVIÇOS GERAIS' and se_count == se_names.index(se) + 1:
      i = 1
      for row in range(1, memo.max_row + 1):
        if memo[f'{conferencia_column}{row}'].value == "Treinamento" or memo[f'{conferencia_column}{row}'].value == "Comissionamento" or memo[f'{conferencia_column}{row}'].value == "Supervisão de Montagem" or memo[f'{conferencia_column}{row}'].value == "Administração de Obra" or memo[f'{conferencia_column}{row}'].value == "Frete" or memo[f'{conferencia_column}{row}'].value == "Despesas de Viagem":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row, row, i)
            i += 1

  make_taxes(se=se, subtopico='SERVIÇOS GERAIS', pis_confins=pis_confins_sv, icms=0, iss=iss_cliente, ipi=ipi)


def make_equipamentos(se):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Demais equipamentos de pátio" or cell.value == "Transformador de Força" or cell.value == "GIS / Módulo Híbrido":
      if memo[f'{subestacao_column}{memo_row}'].value == se and int(memo[f'{qte_column}{memo_row}'].value) > 0:
        exit = False
  print(exit)
  if exit:
    return

  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ENGENHARIA':
      se_count += 1
  
    if planilha_preco[f'B{planilha_preco_row}'].value == 'ELÉTRICA' and se_count == se_names.index(se) + 1:
      i = 1
      planilha_preco.insert_rows(planilha_preco_row + 1, 1)
      
      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}4']
        new_cell = planilha_preco.cell(row=planilha_preco_row + 1, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)

      planilha_preco.cell(row=planilha_preco_row + 1, column=2, value="EQUIPAMENTOS DE PÁTIO")

      for row in range(1, memo.max_row + 1):
        cell = memo[f'{conferencia_column}{row}']
        if cell.value == "Demais equipamentos de pátio" or cell.value == "GIS / Módulo Híbrido" or cell.value == "Transformador de Força":
          if memo[f'{subestacao_column}{row}'].value == se and int(memo[f'{qte_column}{row}'].value) >= 1:
            complete_cells(planilha_preco_row + 1, row, i)
            i += 1
  
  make_taxes(se=se, subtopico='EQUIPAMENTOS DE PÁTIO', pis_confins=pis_confins_eq, icms=icms, iss=0, ipi=ipi)

def make_casa(se):
  exit = True
  for memo_row in range(1, memo.max_row + 1):
    cell = memo[f'{conferencia_column}{memo_row}']
    if cell.value == "Cubículos" or cell.value == "Proteção, medição e controle" or cell.value == "Telecomunicações":
      if memo[f'{subestacao_column}{memo_row}'].value == se and memo[f'{qte_column}{memo_row}'].value >= 1:
        exit = False
  if exit:
    return 0

  se_names = get_se_names()
  se_count = 0
  for planilha_preco_row in range(1, 1000):
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL':
      se_count += 1
    if planilha_preco[f'B{planilha_preco_row}'].value == 'CIVIL' and se_count == se_names.index(se) + 1:
      planilha_preco.insert_rows(planilha_preco_row, 1)

      for planilha_preco_column in range(1, planilha_preco.max_column + 1):
        copy_cell = styles[f'{get_column_letter(planilha_preco_column)}4']
        new_cell = planilha_preco.cell(row=planilha_preco_row, column=planilha_preco_column, value="")
        new_cell._style = copy(copy_cell._style)
      
      planilha_preco.cell(row=planilha_preco_row, column=2, value="CASA DE COMANDO")
  
def make_eletrica(se):
  make_equipamentos(se)
  make_casa(se)


def build():
  se_names = get_se_names()
  make_titles(se_names)
  for se in se_names:
    make_engenharia(se)
    make_eletrica(se)
    make_civil(se, se_names)
    make_montagem(se, se_names)
    make_servicos_gerais(se)

build()
wb_planilha_preco.save('Nova.xlsx')
